#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel CLI - Comprehensive Excel file manipulation.

Read, write, edit, and format Excel files (.xlsx, .xls).

Usage:
    excel.py info <file>
    excel.py read <file> [--sheet NAME] [--range A1:B10] [--format json|csv|markdown]
    excel.py cell <file> <cell> [--sheet NAME]
    excel.py create <file> [--sheets NAME,NAME2]
    excel.py write <file> --data JSON [--sheet NAME] [--start A1]
    excel.py from-csv <csv_file> <excel_file> [--sheet NAME]
    excel.py from-json <json_file> <excel_file> [--sheet NAME]
    excel.py edit <file> <cell> <value> [--sheet NAME] [--formula]
    excel.py add-sheet <file> <name> [--position N]
    excel.py rename-sheet <file> <old_name> <new_name>
    excel.py delete-sheet <file> <name>
    excel.py copy-sheet <file> <source> <new_name>
    excel.py insert-rows <file> <row> [--count N] [--sheet NAME]
    excel.py insert-cols <file> <col> [--count N] [--sheet NAME]
    excel.py delete-rows <file> <row> [--count N] [--sheet NAME]
    excel.py delete-cols <file> <col> [--count N] [--sheet NAME]
    excel.py merge <file> <range> [--sheet NAME]
    excel.py unmerge <file> <range> [--sheet NAME]
    excel.py format <file> <range> [--sheet NAME] [options...]
    excel.py resize <file> [--row N:HEIGHT] [--col A:WIDTH] [--sheet NAME]
    excel.py freeze <file> <cell> [--sheet NAME]
    excel.py find <file> <text> [--sheet NAME]
    excel.py replace <file> <old> <new> [--sheet NAME]
    excel.py to-csv <file> <output> [--sheet NAME]
    excel.py to-json <file> <output> [--sheet NAME]
    excel.py to-markdown <file> [--sheet NAME]
"""

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Optional, Union

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils.cell import coordinate_from_string
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.styles.colors import Color
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


def require_openpyxl():
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


def ok(data: Any):
    """Output success JSON."""
    print(json.dumps({"success": True, **data}, indent=2, default=str))


def fail(message: str, details: Any = None):
    """Output error JSON."""
    result = {"success": False, "error": message}
    if details:
        result["details"] = details
    print(json.dumps(result, indent=2, default=str))
    sys.exit(1)


def parse_range(range_str: str) -> tuple:
    """Parse A1:B10 into ((1,1), (10,2))."""
    if ':' in range_str:
        start, end = range_str.upper().split(':')
    else:
        start = end = range_str.upper()
    
    start_col, start_row = coordinate_from_string(start)
    end_col, end_row = coordinate_from_string(end)
    
    return (
        (start_row, column_index_from_string(start_col)),
        (end_row, column_index_from_string(end_col))
    )


def cell_to_coords(cell: str) -> tuple:
    """Convert A1 to (row, col) tuple."""
    col_str, row = coordinate_from_string(cell.upper())
    return (row, column_index_from_string(col_str))


def coords_to_cell(row: int, col: int) -> str:
    """Convert (row, col) to A1 notation."""
    return f"{get_column_letter(col)}{row}"


def get_sheet(wb, sheet_name: Optional[str] = None):
    """Get worksheet by name or active sheet."""
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            fail(f"Sheet '{sheet_name}' not found", {"available": wb.sheetnames})
        return wb[sheet_name]
    return wb.active


def read_sheet_data(ws, range_str: Optional[str] = None) -> list:
    """Read sheet data as list of lists."""
    if range_str:
        (start_row, start_col), (end_row, end_col) = parse_range(range_str)
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            data.append(row_data)
        return data
    else:
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        return data


def data_to_markdown(data: list, headers: bool = True) -> str:
    """Convert data to markdown table."""
    if not data:
        return "*Empty*"
    
    # Calculate column widths
    widths = []
    for col_idx in range(len(data[0])):
        max_width = 3
        for row in data:
            if col_idx < len(row):
                val = str(row[col_idx]) if row[col_idx] is not None else ""
                max_width = max(max_width, len(val))
        widths.append(max_width)
    
    lines = []
    for i, row in enumerate(data):
        cells = []
        for j, val in enumerate(row):
            val_str = str(val) if val is not None else ""
            cells.append(val_str.ljust(widths[j]))
        lines.append("| " + " | ".join(cells) + " |")
        
        if i == 0 and headers:
            sep = ["-" * w for w in widths]
            lines.append("| " + " | ".join(sep) + " |")
    
    return "\n".join(lines)


def parse_color(color_str: str) -> str:
    """Parse color string to ARGB hex."""
    color_str = color_str.upper().strip()
    
    # Named colors
    colors = {
        "RED": "FFFF0000",
        "GREEN": "FF00FF00",
        "BLUE": "FF0000FF",
        "YELLOW": "FFFFFF00",
        "WHITE": "FFFFFFFF",
        "BLACK": "FF000000",
        "GRAY": "FF808080",
        "GREY": "FF808080",
        "ORANGE": "FFFFA500",
        "PURPLE": "FF800080",
        "PINK": "FFFFC0CB",
        "CYAN": "FF00FFFF",
    }
    
    if color_str in colors:
        return colors[color_str]
    
    # Hex color (with or without #)
    if color_str.startswith("#"):
        color_str = color_str[1:]
    
    if len(color_str) == 6:
        return "FF" + color_str
    elif len(color_str) == 8:
        return color_str
    
    return "FF000000"  # Default black


# ============================================================================
# Commands
# ============================================================================

def cmd_info(args):
    """Get workbook information."""
    require_openpyxl()
    
    try:
        wb = load_workbook(args.file, read_only=False, data_only=True)
    except Exception as e:
        fail(f"Failed to open file: {e}")
    
    sheets_info = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets_info.append({
            "name": name,
            "dimensions": ws.dimensions or "A1",
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        })
    
    ok({
        "file": args.file,
        "sheets": sheets_info,
        "sheet_count": len(wb.sheetnames),
        "active_sheet": wb.active.title if wb.active else None,
    })
    wb.close()


def cmd_read(args):
    """Read sheet data."""
    require_openpyxl()
    
    try:
        wb = load_workbook(args.file, read_only=True, data_only=True)
    except Exception as e:
        fail(f"Failed to open file: {e}")
    
    ws = get_sheet(wb, args.sheet)
    data = read_sheet_data(ws, args.range)
    
    if args.format == "csv":
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        for row in data:
            writer.writerow(row)
        print(output.getvalue())
    elif args.format == "markdown":
        print(data_to_markdown(data))
    else:  # json
        ok({
            "sheet": ws.title,
            "range": args.range or ws.dimensions,
            "rows": len(data),
            "columns": len(data[0]) if data else 0,
            "data": data,
        })
    
    wb.close()


def cmd_cell(args):
    """Read a specific cell."""
    require_openpyxl()
    
    try:
        wb = load_workbook(args.file, read_only=True, data_only=False)
    except Exception as e:
        fail(f"Failed to open file: {e}")
    
    ws = get_sheet(wb, args.sheet)
    cell = ws[args.cell.upper()]
    
    ok({
        "cell": args.cell.upper(),
        "value": cell.value,
        "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None,
        "data_type": cell.data_type,
        "is_merged": cell.coordinate in ws.merged_cells,
    })
    wb.close()


def cmd_create(args):
    """Create a new workbook."""
    require_openpyxl()
    
    wb = Workbook()
    
    if args.sheets:
        sheet_names = [s.strip() for s in args.sheets.split(",")]
        # Rename first sheet
        wb.active.title = sheet_names[0]
        # Add additional sheets
        for name in sheet_names[1:]:
            wb.create_sheet(title=name)
    
    try:
        wb.save(args.file)
    except Exception as e:
        fail(f"Failed to save file: {e}")
    
    ok({
        "file": args.file,
        "sheets": wb.sheetnames,
        "created": True,
    })


def cmd_write(args):
    """Write data to cells."""
    require_openpyxl()
    
    # Load or create workbook
    if os.path.exists(args.file):
        wb = load_workbook(args.file)
    else:
        wb = Workbook()
    
    ws = get_sheet(wb, args.sheet)
    
    # Parse data
    try:
        data = json.loads(args.data)
    except json.JSONDecodeError as e:
        fail(f"Invalid JSON data: {e}")
    
    # Parse start cell
    start_row, start_col = cell_to_coords(args.start)
    
    # Write data
    if isinstance(data, list):
        for i, row in enumerate(data):
            if isinstance(row, list):
                for j, val in enumerate(row):
                    ws.cell(row=start_row + i, column=start_col + j, value=val)
            else:
                ws.cell(row=start_row + i, column=start_col, value=row)
    elif isinstance(data, dict):
        # Write as key-value pairs or headers + rows
        if "headers" in data and "rows" in data:
            for j, header in enumerate(data["headers"]):
                ws.cell(row=start_row, column=start_col + j, value=header)
            for i, row in enumerate(data["rows"]):
                for j, val in enumerate(row):
                    ws.cell(row=start_row + 1 + i, column=start_col + j, value=val)
        else:
            for i, (key, val) in enumerate(data.items()):
                ws.cell(row=start_row + i, column=start_col, value=key)
                ws.cell(row=start_row + i, column=start_col + 1, value=val)
    else:
        ws.cell(row=start_row, column=start_col, value=data)
    
    wb.save(args.file)
    ok({
        "file": args.file,
        "sheet": ws.title,
        "start": args.start,
        "written": True,
    })


def cmd_from_csv(args):
    """Create Excel from CSV."""
    require_openpyxl()
    
    wb = Workbook()
    ws = wb.active
    if args.sheet:
        ws.title = args.sheet
    
    try:
        with open(args.csv_file, 'r', newline='', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
    except Exception as e:
        fail(f"Failed to read CSV: {e}")
    
    wb.save(args.excel_file)
    ok({
        "source": args.csv_file,
        "output": args.excel_file,
        "sheet": ws.title,
        "rows": ws.max_row,
    })


def cmd_from_json(args):
    """Create Excel from JSON."""
    require_openpyxl()
    
    try:
        with open(args.json_file, 'r') as f:
            data = json.load(f)
    except Exception as e:
        fail(f"Failed to read JSON: {e}")
    
    wb = Workbook()
    ws = wb.active
    if args.sheet:
        ws.title = args.sheet
    
    if isinstance(data, list):
        if data and isinstance(data[0], dict):
            # List of dicts - use keys as headers
            headers = list(data[0].keys())
            ws.append(headers)
            for item in data:
                ws.append([item.get(h) for h in headers])
        else:
            # List of lists or values
            for row in data:
                if isinstance(row, list):
                    ws.append(row)
                else:
                    ws.append([row])
    elif isinstance(data, dict):
        if "headers" in data and "rows" in data:
            ws.append(data["headers"])
            for row in data["rows"]:
                ws.append(row)
        else:
            for key, val in data.items():
                ws.append([key, val])
    
    wb.save(args.excel_file)
    ok({
        "source": args.json_file,
        "output": args.excel_file,
        "sheet": ws.title,
        "rows": ws.max_row,
    })


def cmd_edit(args):
    """Edit a cell value or formula."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    cell = ws[args.cell.upper()]
    old_value = cell.value
    
    # Set new value
    if args.formula:
        if not args.value.startswith("="):
            args.value = "=" + args.value
        cell.value = args.value
    else:
        # Try to convert to appropriate type
        try:
            if args.value.lower() in ('true', 'false'):
                cell.value = args.value.lower() == 'true'
            elif '.' in args.value:
                cell.value = float(args.value)
            else:
                cell.value = int(args.value)
        except ValueError:
            cell.value = args.value
    
    wb.save(args.file)
    ok({
        "cell": args.cell.upper(),
        "old_value": old_value,
        "new_value": cell.value,
        "is_formula": args.formula,
    })


def cmd_add_sheet(args):
    """Add a new sheet."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    
    if args.name in wb.sheetnames:
        fail(f"Sheet '{args.name}' already exists")
    
    if args.position is not None:
        wb.create_sheet(title=args.name, index=args.position)
    else:
        wb.create_sheet(title=args.name)
    
    wb.save(args.file)
    ok({
        "added": args.name,
        "sheets": wb.sheetnames,
    })


def cmd_rename_sheet(args):
    """Rename a sheet."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    
    if args.old_name not in wb.sheetnames:
        fail(f"Sheet '{args.old_name}' not found")
    
    if args.new_name in wb.sheetnames:
        fail(f"Sheet '{args.new_name}' already exists")
    
    wb[args.old_name].title = args.new_name
    wb.save(args.file)
    
    ok({
        "old_name": args.old_name,
        "new_name": args.new_name,
        "sheets": wb.sheetnames,
    })


def cmd_delete_sheet(args):
    """Delete a sheet."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    
    if args.name not in wb.sheetnames:
        fail(f"Sheet '{args.name}' not found")
    
    if len(wb.sheetnames) == 1:
        fail("Cannot delete the only sheet in workbook")
    
    del wb[args.name]
    wb.save(args.file)
    
    ok({
        "deleted": args.name,
        "sheets": wb.sheetnames,
    })


def cmd_copy_sheet(args):
    """Copy a sheet."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    
    if args.source not in wb.sheetnames:
        fail(f"Sheet '{args.source}' not found")
    
    if args.new_name in wb.sheetnames:
        fail(f"Sheet '{args.new_name}' already exists")
    
    source_ws = wb[args.source]
    new_ws = wb.copy_worksheet(source_ws)
    new_ws.title = args.new_name
    
    wb.save(args.file)
    ok({
        "source": args.source,
        "copy": args.new_name,
        "sheets": wb.sheetnames,
    })


def cmd_insert_rows(args):
    """Insert rows."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.insert_rows(args.row, args.count)
    wb.save(args.file)
    
    ok({
        "inserted": "rows",
        "at": args.row,
        "count": args.count,
        "sheet": ws.title,
    })


def cmd_insert_cols(args):
    """Insert columns."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    # Convert column letter to number if needed
    col = args.col
    if isinstance(col, str) and col.isalpha():
        col = column_index_from_string(col.upper())
    
    ws.insert_cols(int(col), args.count)
    wb.save(args.file)
    
    ok({
        "inserted": "columns",
        "at": args.col,
        "count": args.count,
        "sheet": ws.title,
    })


def cmd_delete_rows(args):
    """Delete rows."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.delete_rows(args.row, args.count)
    wb.save(args.file)
    
    ok({
        "deleted": "rows",
        "at": args.row,
        "count": args.count,
        "sheet": ws.title,
    })


def cmd_delete_cols(args):
    """Delete columns."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    col = args.col
    if isinstance(col, str) and col.isalpha():
        col = column_index_from_string(col.upper())
    
    ws.delete_cols(int(col), args.count)
    wb.save(args.file)
    
    ok({
        "deleted": "columns",
        "at": args.col,
        "count": args.count,
        "sheet": ws.title,
    })


def cmd_merge(args):
    """Merge cells."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.merge_cells(args.range)
    wb.save(args.file)
    
    ok({
        "merged": args.range,
        "sheet": ws.title,
    })


def cmd_unmerge(args):
    """Unmerge cells."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.unmerge_cells(args.range)
    wb.save(args.file)
    
    ok({
        "unmerged": args.range,
        "sheet": ws.title,
    })


def cmd_format(args):
    """Format cells."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    (start_row, start_col), (end_row, end_col) = parse_range(args.range)
    
    applied = []
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            
            # Font
            if args.bold is not None or args.italic is not None or args.font_size or args.font_color or args.font_name:
                current_font = cell.font
                cell.font = Font(
                    name=args.font_name or current_font.name,
                    size=args.font_size or current_font.size,
                    bold=args.bold if args.bold is not None else current_font.bold,
                    italic=args.italic if args.italic is not None else current_font.italic,
                    color=parse_color(args.font_color) if args.font_color else current_font.color,
                )
                applied.append("font")
            
            # Fill/background
            if args.bg_color:
                cell.fill = PatternFill(start_color=parse_color(args.bg_color),
                                         end_color=parse_color(args.bg_color),
                                         fill_type="solid")
                applied.append("background")
            
            # Alignment
            if args.align or args.valign or args.wrap:
                current_align = cell.alignment
                cell.alignment = Alignment(
                    horizontal=args.align or current_align.horizontal,
                    vertical=args.valign or current_align.vertical,
                    wrap_text=args.wrap if args.wrap is not None else current_align.wrap_text,
                )
                applied.append("alignment")
            
            # Border
            if args.border:
                side = Side(style=args.border, color="FF000000")
                cell.border = Border(left=side, right=side, top=side, bottom=side)
                applied.append("border")
    
    wb.save(args.file)
    ok({
        "range": args.range,
        "sheet": ws.title,
        "applied": list(set(applied)),
    })


def cmd_resize(args):
    """Resize rows/columns."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    resized = []
    
    if args.row:
        for spec in args.row:
            row_num, height = spec.split(':')
            ws.row_dimensions[int(row_num)].height = float(height)
            resized.append(f"row {row_num} = {height}")
    
    if args.col:
        for spec in args.col:
            col_letter, width = spec.split(':')
            ws.column_dimensions[col_letter.upper()].width = float(width)
            resized.append(f"col {col_letter} = {width}")
    
    wb.save(args.file)
    ok({
        "sheet": ws.title,
        "resized": resized,
    })


def cmd_freeze(args):
    """Freeze panes at cell."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    ws.freeze_panes = args.cell.upper()
    wb.save(args.file)
    
    ok({
        "sheet": ws.title,
        "frozen_at": args.cell.upper(),
    })


def cmd_find(args):
    """Find text in sheet."""
    require_openpyxl()
    
    wb = load_workbook(args.file, read_only=True, data_only=True)
    ws = get_sheet(wb, args.sheet)
    
    results = []
    search_lower = args.text.lower()
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and search_lower in str(cell.value).lower():
                results.append({
                    "cell": cell.coordinate,
                    "value": cell.value,
                })
    
    ok({
        "search": args.text,
        "sheet": ws.title,
        "found": len(results),
        "results": results,
    })
    wb.close()


def cmd_replace(args):
    """Find and replace text."""
    require_openpyxl()
    
    wb = load_workbook(args.file)
    ws = get_sheet(wb, args.sheet)
    
    count = 0
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and args.old in cell.value:
                cell.value = cell.value.replace(args.old, args.new)
                count += 1
    
    wb.save(args.file)
    ok({
        "old": args.old,
        "new": args.new,
        "sheet": ws.title,
        "replaced": count,
    })


def cmd_to_csv(args):
    """Export sheet to CSV."""
    require_openpyxl()
    
    wb = load_workbook(args.file, read_only=True, data_only=True)
    ws = get_sheet(wb, args.sheet)
    
    with open(args.output, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    
    ok({
        "source": args.file,
        "sheet": ws.title,
        "output": args.output,
        "rows": ws.max_row,
    })
    wb.close()


def cmd_to_json(args):
    """Export sheet to JSON."""
    require_openpyxl()
    
    wb = load_workbook(args.file, read_only=True, data_only=True)
    ws = get_sheet(wb, args.sheet)
    
    data = []
    headers = None
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = list(row)
        else:
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j]:
                    row_dict[headers[j]] = val
            data.append(row_dict)
    
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, default=str)
    
    ok({
        "source": args.file,
        "sheet": ws.title,
        "output": args.output,
        "rows": len(data),
    })
    wb.close()


def cmd_to_markdown(args):
    """Export sheet to markdown table."""
    require_openpyxl()
    
    wb = load_workbook(args.file, read_only=True, data_only=True)
    ws = get_sheet(wb, args.sheet)
    
    data = read_sheet_data(ws)
    print(data_to_markdown(data))
    wb.close()


def main():
    parser = argparse.ArgumentParser(
        description="Excel CLI - Read, write, edit, and format Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    subparsers = parser.add_subparsers(dest="command", help="Commands")
    
    # info
    p = subparsers.add_parser("info", help="Get workbook information")
    p.add_argument("file", help="Excel file path")
    
    # read
    p = subparsers.add_parser("read", help="Read sheet data")
    p.add_argument("file", help="Excel file path")
    p.add_argument("--sheet", "-s", help="Sheet name")
    p.add_argument("--range", "-r", help="Cell range (e.g., A1:D10)")
    p.add_argument("--format", "-f", choices=["json", "csv", "markdown"], default="json")
    
    # cell
    p = subparsers.add_parser("cell", help="Read a specific cell")
    p.add_argument("file", help="Excel file path")
    p.add_argument("cell", help="Cell reference (e.g., A1)")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # create
    p = subparsers.add_parser("create", help="Create new workbook")
    p.add_argument("file", help="Output file path")
    p.add_argument("--sheets", help="Comma-separated sheet names")
    
    # write
    p = subparsers.add_parser("write", help="Write data to cells")
    p.add_argument("file", help="Excel file path")
    p.add_argument("--data", "-d", required=True, help="JSON data to write")
    p.add_argument("--sheet", "-s", help="Sheet name")
    p.add_argument("--start", default="A1", help="Starting cell (default: A1)")
    
    # from-csv
    p = subparsers.add_parser("from-csv", help="Create Excel from CSV")
    p.add_argument("csv_file", help="Input CSV file")
    p.add_argument("excel_file", help="Output Excel file")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # from-json
    p = subparsers.add_parser("from-json", help="Create Excel from JSON")
    p.add_argument("json_file", help="Input JSON file")
    p.add_argument("excel_file", help="Output Excel file")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # edit
    p = subparsers.add_parser("edit", help="Edit a cell")
    p.add_argument("file", help="Excel file path")
    p.add_argument("cell", help="Cell reference")
    p.add_argument("value", help="New value")
    p.add_argument("--sheet", "-s", help="Sheet name")
    p.add_argument("--formula", "-F", action="store_true", help="Value is a formula")
    
    # add-sheet
    p = subparsers.add_parser("add-sheet", help="Add a new sheet")
    p.add_argument("file", help="Excel file path")
    p.add_argument("name", help="New sheet name")
    p.add_argument("--position", "-p", type=int, help="Position index")
    
    # rename-sheet
    p = subparsers.add_parser("rename-sheet", help="Rename a sheet")
    p.add_argument("file", help="Excel file path")
    p.add_argument("old_name", help="Current sheet name")
    p.add_argument("new_name", help="New sheet name")
    
    # delete-sheet
    p = subparsers.add_parser("delete-sheet", help="Delete a sheet")
    p.add_argument("file", help="Excel file path")
    p.add_argument("name", help="Sheet name to delete")
    
    # copy-sheet
    p = subparsers.add_parser("copy-sheet", help="Copy a sheet")
    p.add_argument("file", help="Excel file path")
    p.add_argument("source", help="Source sheet name")
    p.add_argument("new_name", help="New sheet name")
    
    # insert-rows
    p = subparsers.add_parser("insert-rows", help="Insert rows")
    p.add_argument("file", help="Excel file path")
    p.add_argument("row", type=int, help="Row number to insert at")
    p.add_argument("--count", "-n", type=int, default=1, help="Number of rows")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # insert-cols
    p = subparsers.add_parser("insert-cols", help="Insert columns")
    p.add_argument("file", help="Excel file path")
    p.add_argument("col", help="Column (letter or number) to insert at")
    p.add_argument("--count", "-n", type=int, default=1, help="Number of columns")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # delete-rows
    p = subparsers.add_parser("delete-rows", help="Delete rows")
    p.add_argument("file", help="Excel file path")
    p.add_argument("row", type=int, help="Starting row number")
    p.add_argument("--count", "-n", type=int, default=1, help="Number of rows")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # delete-cols
    p = subparsers.add_parser("delete-cols", help="Delete columns")
    p.add_argument("file", help="Excel file path")
    p.add_argument("col", help="Column (letter or number)")
    p.add_argument("--count", "-n", type=int, default=1, help="Number of columns")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # merge
    p = subparsers.add_parser("merge", help="Merge cells")
    p.add_argument("file", help="Excel file path")
    p.add_argument("range", help="Cell range to merge (e.g., A1:C1)")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # unmerge
    p = subparsers.add_parser("unmerge", help="Unmerge cells")
    p.add_argument("file", help="Excel file path")
    p.add_argument("range", help="Cell range to unmerge")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # format
    p = subparsers.add_parser("format", help="Format cells")
    p.add_argument("file", help="Excel file path")
    p.add_argument("range", help="Cell range to format")
    p.add_argument("--sheet", "-s", help="Sheet name")
    p.add_argument("--bold", "-b", action="store_true", default=None)
    p.add_argument("--no-bold", dest="bold", action="store_false")
    p.add_argument("--italic", "-i", action="store_true", default=None)
    p.add_argument("--no-italic", dest="italic", action="store_false")
    p.add_argument("--font-size", type=int, help="Font size")
    p.add_argument("--font-color", help="Font color (name or hex)")
    p.add_argument("--font-name", help="Font name")
    p.add_argument("--bg-color", help="Background color")
    p.add_argument("--align", choices=["left", "center", "right"])
    p.add_argument("--valign", choices=["top", "center", "bottom"])
    p.add_argument("--wrap", action="store_true", default=None)
    p.add_argument("--no-wrap", dest="wrap", action="store_false")
    p.add_argument("--border", choices=["thin", "medium", "thick", "double"])
    
    # resize
    p = subparsers.add_parser("resize", help="Resize rows/columns")
    p.add_argument("file", help="Excel file path")
    p.add_argument("--row", action="append", help="Row:height (e.g., 1:30)")
    p.add_argument("--col", action="append", help="Col:width (e.g., A:20)")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # freeze
    p = subparsers.add_parser("freeze", help="Freeze panes")
    p.add_argument("file", help="Excel file path")
    p.add_argument("cell", help="Cell to freeze at (e.g., A2 freezes row 1)")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # find
    p = subparsers.add_parser("find", help="Find text in sheet")
    p.add_argument("file", help="Excel file path")
    p.add_argument("text", help="Text to search for")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # replace
    p = subparsers.add_parser("replace", help="Find and replace text")
    p.add_argument("file", help="Excel file path")
    p.add_argument("old", help="Text to find")
    p.add_argument("new", help="Replacement text")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # to-csv
    p = subparsers.add_parser("to-csv", help="Export sheet to CSV")
    p.add_argument("file", help="Excel file path")
    p.add_argument("output", help="Output CSV file")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # to-json
    p = subparsers.add_parser("to-json", help="Export sheet to JSON")
    p.add_argument("file", help="Excel file path")
    p.add_argument("output", help="Output JSON file")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    # to-markdown
    p = subparsers.add_parser("to-markdown", help="Export sheet to markdown")
    p.add_argument("file", help="Excel file path")
    p.add_argument("--sheet", "-s", help="Sheet name")
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        sys.exit(1)
    
    commands = {
        "info": cmd_info,
        "read": cmd_read,
        "cell": cmd_cell,
        "create": cmd_create,
        "write": cmd_write,
        "from-csv": cmd_from_csv,
        "from-json": cmd_from_json,
        "edit": cmd_edit,
        "add-sheet": cmd_add_sheet,
        "rename-sheet": cmd_rename_sheet,
        "delete-sheet": cmd_delete_sheet,
        "copy-sheet": cmd_copy_sheet,
        "insert-rows": cmd_insert_rows,
        "insert-cols": cmd_insert_cols,
        "delete-rows": cmd_delete_rows,
        "delete-cols": cmd_delete_cols,
        "merge": cmd_merge,
        "unmerge": cmd_unmerge,
        "format": cmd_format,
        "resize": cmd_resize,
        "freeze": cmd_freeze,
        "find": cmd_find,
        "replace": cmd_replace,
        "to-csv": cmd_to_csv,
        "to-json": cmd_to_json,
        "to-markdown": cmd_to_markdown,
    }
    
    try:
        commands[args.command](args)
    except Exception as e:
        fail(str(e))


if __name__ == "__main__":
    main()
